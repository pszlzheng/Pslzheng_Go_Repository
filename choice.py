import os
import sys
import openpyxl
import xlrd
from pymysql import *
from datetime import *
import matplotlib.pyplot as plt
import serial.tools.list_ports
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtCore import QTimer
from PyQt5 import QtCore, QtGui, QtWidgets
import pymysql

device={'STM_ST':0,'STM_ND':1,'STM_RD':2}
fish_name={0:'罗非',1:'青鱼'}
fish_age={0:'1月龄',1:'3月龄',2:'5月龄'}
pool_ph={0:'7.0',1:'7.1',2:'7.2',3:'7.3',4:'7.4',5:'7.5',6:'7.6',7:'7.7',8:'7.8',9:'7.9'}
pool_oxy={0:'0.18',1:'0.19',2:'0.20',3:'0.21',4:'0.22'}
pool_temp={0:'18',1:'20',2:'22',3:'24',4:'26',5:'28',6:'30',7:'32',8:'34',9:'36'}

dt=datetime.now()
day=str(dt.year)+'-'+str(dt.month)+'-'+str(dt.day)
#minute=str(dt.year)+'-'+str(dt.month)+'-'+str(dt.day)+'-'+str(dt.hour)+'-'+str(dt.minute)
minute=str(dt.hour)+'-'+str(dt.minute)

class Ui_Form(object):

    def setupUi(self, Form):
        Form.setObjectName("Format")
        Form.setFixedSize(700, 460)

        #串口配置框
        self.formGroupBox = QtWidgets.QGroupBox(Form)
        self.formGroupBox.setGeometry(QtCore.QRect(20, 20, 170, 300))
        self.formGroupBox.setObjectName("formGroupBox")
        self.formLayout_1 = QtWidgets.QFormLayout(self.formGroupBox)
        self.formLayout_1.setContentsMargins(10, 10, 10, 10)
        self.formLayout_1.setSpacing(10)
        self.formLayout_1.setObjectName("formLayout_1")
        self.s1__lb_1 = QtWidgets.QLabel(self.formGroupBox)
        self.s1__lb_1.setObjectName("s1__lb_1")
        self.formLayout_1.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.s1__lb_1)
        self.s1__box_1 = QtWidgets.QPushButton(self.formGroupBox)
        self.s1__box_1.setAutoRepeatInterval(100)
        self.s1__box_1.setDefault(True)
        self.s1__box_1.setObjectName("s1__box_1")
        self.formLayout_1.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.s1__box_1)
        self.s1__lb_2 = QtWidgets.QLabel(self.formGroupBox)
        self.s1__lb_2.setObjectName("s1__lb_2")
        self.formLayout_1.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.s1__lb_2)
        self.s1__box_2 = QtWidgets.QComboBox(self.formGroupBox)
        self.s1__box_2.setObjectName("s1__box_2")
        self.formLayout_1.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.s1__box_2)
        self.s1__lb_3 = QtWidgets.QLabel(self.formGroupBox)
        self.s1__lb_3.setObjectName("s1__lb_3")
        self.formLayout_1.setWidget(3, QtWidgets.QFormLayout.LabelRole, self.s1__lb_3)
        self.s1__box_3 = QtWidgets.QComboBox(self.formGroupBox)
        self.s1__box_3.setObjectName("s1__box_3")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.formLayout_1.setWidget(3,QtWidgets.QFormLayout.FieldRole,self.s1__box_3)
        self.s1__lb_4 = QtWidgets.QLabel(self.formGroupBox)
        self.s1__lb_4.setObjectName("s1__lb_4")
        self.formLayout_1.setWidget(4, QtWidgets.QFormLayout.LabelRole, self.s1__lb_4)
        self.s1__box_4 = QtWidgets.QComboBox(self.formGroupBox)
        self.s1__box_4.setObjectName("s1__box_4")
        self.s1__box_4.addItem("")
        self.s1__box_4.addItem("")
        self.s1__box_4.addItem("")
        self.s1__box_4.addItem("")
        self.formLayout_1.setWidget(4, QtWidgets.QFormLayout.FieldRole, self.s1__box_4)
        self.s1__lb_5 = QtWidgets.QLabel(self.formGroupBox)
        self.s1__lb_5.setObjectName("s1__lb_5")
        self.formLayout_1.setWidget(5, QtWidgets.QFormLayout.LabelRole, self.s1__lb_5)
        self.s1__box_5 = QtWidgets.QComboBox(self.formGroupBox)
        self.s1__box_5.setObjectName("s1__box_5")
        self.s1__box_5.addItem("")
        self.formLayout_1.setWidget(5, QtWidgets.QFormLayout.FieldRole, self.s1__box_5)
        self.open_button = QtWidgets.QPushButton(self.formGroupBox)
        self.open_button.setObjectName("open_button")
        self.formLayout_1.setWidget(7, QtWidgets.QFormLayout.SpanningRole, self.open_button)
        self.close_button = QtWidgets.QPushButton(self.formGroupBox)
        self.close_button.setObjectName("close_button")
        self.formLayout_1.setWidget(8, QtWidgets.QFormLayout.SpanningRole, self.close_button)
        self.s1__lb_6 = QtWidgets.QLabel(self.formGroupBox)
        self.s1__lb_6.setObjectName("s1__lb_6")
        self.formLayout_1.setWidget(6, QtWidgets.QFormLayout.LabelRole, self.s1__lb_6)
        self.s1__box_6 = QtWidgets.QComboBox(self.formGroupBox)
        self.s1__box_6.setObjectName("s1__box_6")
        self.s1__box_6.addItem("")
        self.formLayout_1.setWidget(6, QtWidgets.QFormLayout.FieldRole, self.s1__box_6)
        self.state_label = QtWidgets.QLabel(self.formGroupBox)
        self.state_label.setText("")
        self.state_label.setTextFormat(QtCore.Qt.AutoText)
        self.state_label.setScaledContents(True)
        self.state_label.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.state_label.setObjectName("state_label")
        self.formLayout_1.setWidget(2, QtWidgets.QFormLayout.SpanningRole, self.state_label)
        
        #消息接收框
        self.verticalGroupBox_1 = QtWidgets.QGroupBox(Form)
        self.verticalGroupBox_1.setGeometry(QtCore.QRect(210, 20, 400, 100))
        self.verticalGroupBox_1.setObjectName("verticalGroupBox")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalGroupBox_1)
        self.verticalLayout.setContentsMargins(10, 10, 10, 10)
        self.verticalLayout.setObjectName("verticalLayout")

        self.s2__receive_text = QtWidgets.QTextBrowser(self.verticalGroupBox_1)
        self.s2__receive_text.setObjectName("s2__receive_text")
        self.verticalLayout.addWidget(self.s2__receive_text)
        
        #消息发送框
        self.verticalGroupBox_2 = QtWidgets.QGroupBox(Form)
        self.verticalGroupBox_2.setGeometry(QtCore.QRect(210, 300, 470, 65))
        self.verticalGroupBox_2.setObjectName("verticalGroupBox_2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalGroupBox_2)
        self.verticalLayout_2.setContentsMargins(10, 10, 10, 10)
        self.verticalLayout_2.setObjectName("verticalLayout_2") 

        self.s3__send_text = QtWidgets.QTextEdit(self.verticalGroupBox_2)
        self.s3__send_text.setObjectName("s3__send_text")
        self.verticalLayout_2.addWidget(self.s3__send_text)
        self.s3__send_button = QtWidgets.QPushButton(Form)
        self.s3__send_button.setGeometry(QtCore.QRect(210, 370, 110, 30))
        self.s3__send_button.setObjectName("s3__send_button")
        self.s3__clear_button = QtWidgets.QPushButton(Form)
        self.s3__clear_button.setGeometry(QtCore.QRect(395, 370, 110, 30))
        self.s3__clear_button.setObjectName("s3__clear_button")
        self.s3__check_button = QtWidgets.QPushButton(Form)
        self.s3__check_button.setGeometry(QtCore.QRect(570, 370, 110, 30))
        self.s3__check_button.setObjectName("s3__check_button")

        #信息查询框
        self.verticalGroupBox_3=QtWidgets.QGroupBox(Form)
        self.verticalGroupBox_3.setGeometry(QtCore.QRect(210,130,220,160))
        self.verticalGroupBox_3.setObjectName("verticalGroupBox_3")
        self.verticalLayout_3=QtWidgets.QVBoxLayout(self.verticalGroupBox_3)
        self.verticalLayout_3.setContentsMargins(10,10,10,10)
        self.verticalLayout_3.setObjectName('verticalLayout_3')

        self.s4__PH_lab=QtWidgets.QLabel(Form)
        self.s4__PH_lab.setObjectName('s4_PH_lab')
        self.s4__PH_lab.setGeometry(QtCore.QRect(240,150,60,30))

        self.s4__PH_val=QtWidgets.QLabel(Form)
        self.s4__PH_val.setObjectName('s4_PH_val')
        self.s4__PH_val.setGeometry(QtCore.QRect(320,150,60,30))

        self.s4__OXY_lab=QtWidgets.QLabel(Form)
        self.s4__OXY_lab.setObjectName('s4_OXY_lab')
        self.s4__OXY_lab.setGeometry(QtCore.QRect(240,200,60,30))

        self.s4__OXY_val=QtWidgets.QLabel(Form)
        self.s4__OXY_val.setObjectName('s4_OXY_val')
        self.s4__OXY_val.setGeometry(QtCore.QRect(320,200,60,30))

        self.s4__TEMP_lab=QtWidgets.QLabel(Form)
        self.s4__TEMP_lab.setObjectName('s4_TEMP_lab')
        self.s4__TEMP_lab.setGeometry(QtCore.QRect(240,250,60,30))

        self.s4__TEMP_val=QtWidgets.QLabel(Form)
        self.s4__TEMP_val.setObjectName('s4_TEMP_val')
        self.s4__TEMP_val.setGeometry(QtCore.QRect(320,250,60,30))

        self.verticalGroupBox_4=QtWidgets.QGroupBox(Form)
        self.verticalGroupBox_4.setObjectName('verticalGroupBox_4')
        self.verticalGroupBox_4.setGeometry(QtCore.QRect(440,130,280,160))

        self.s4_dt_btn = QtWidgets.QPushButton(Form)#获取用户输入的日期
        self.s4_dt_btn.setGeometry(QtCore.QRect(460, 175, 90, 30))
        self.s4_dt_btn.setObjectName("s4_dt_btn")

        self.s4_lin_day=QtWidgets.QLineEdit('2020-6-11',self)
        self.s4_lin_day.setGeometry(QtCore.QRect(580,175,90,30))

        self.s4_fish_btn = QtWidgets.QPushButton(Form)#获取用户输入的鱼种
        self.s4_fish_btn.setGeometry(QtCore.QRect(460, 140, 90, 30))
        self.s4_fish_btn.setObjectName("s4_dt_btn")

        self.s4_lin_fish=QtWidgets.QLineEdit('青鱼',self)
        self.s4_lin_fish.setGeometry(QtCore.QRect(580,140,90,30))

        self.s4__PH_btn = QtWidgets.QPushButton(Form)#显示PH曲线
        self.s4__PH_btn.setGeometry(QtCore.QRect(460, 210, 90, 30))
        self.s4__PH_btn.setObjectName("s4__PH_text")

        self.s4__TEMP_btn = QtWidgets.QPushButton(Form)#显示温度曲线
        self.s4__TEMP_btn.setGeometry(QtCore.QRect(460, 245, 90, 30))
        self.s4__TEMP_btn.setObjectName("s4__TEMP_text")

        self.s4__OXY_btn = QtWidgets.QPushButton(Form)#显示溶氧量曲线
        self.s4__OXY_btn.setGeometry(QtCore.QRect(580, 210, 90, 30))
        self.s4__OXY_btn.setObjectName("s4__OXY_text")

        self.formGroupBox1 = QtWidgets.QGroupBox(Form)
        self.formGroupBox1.setGeometry(QtCore.QRect(20, 340, 170, 100))
        self.formGroupBox1.setObjectName("formGroupBox1")
        self.formLayout_2 = QtWidgets.QFormLayout(self.formGroupBox1)
        self.formLayout_2.setContentsMargins(10, 10, 10, 10)
        self.formLayout_2.setSpacing(10)
        self.formLayout_2.setObjectName("formLayout_2")
        self.label = QtWidgets.QLabel(self.formGroupBox1)
        self.label.setObjectName("label")
        self.formLayout_2.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.label)
        self.label_2 = QtWidgets.QLabel(self.formGroupBox1)
        self.label_2.setObjectName("label_2")
        self.formLayout_2.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.label_2)
        self.lineEdit = QtWidgets.QLineEdit(self.formGroupBox1)
        self.lineEdit.setObjectName("lineEdit")
        self.formLayout_2.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.lineEdit)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.formGroupBox1)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.formLayout_2.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.lineEdit_2)
        self.hex_receive = QtWidgets.QCheckBox(Form)
        self.hex_receive.setGeometry(QtCore.QRect(620, 40, 71, 16))
        self.hex_receive.setObjectName("hex_receive")
        self.s2__clear_button = QtWidgets.QPushButton(Form)
        self.s2__clear_button.setGeometry(QtCore.QRect(620, 80, 61, 31))
        self.s2__clear_button.setObjectName("s2__clear_button")
        self.timer_send_cb = QtWidgets.QCheckBox(Form)
        self.timer_send_cb.setGeometry(QtCore.QRect(260, 420, 71, 16))
        self.timer_send_cb.setObjectName("timer_send_cb")
        self.lineEdit_3 = QtWidgets.QLineEdit(Form)
        self.lineEdit_3.setGeometry(QtCore.QRect(350, 420, 61, 20))
        self.lineEdit_3.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.dw = QtWidgets.QLabel(Form)
        self.dw.setGeometry(QtCore.QRect(420, 420, 54, 20))
        self.dw.setObjectName("dw")

        #最上层显示
        self.verticalGroupBox_1.raise_()
        self.verticalGroupBox_2.raise_()
        self.formGroupBox.raise_()
        self.s3__send_button.raise_()
        self.s3__clear_button.raise_()
        self.s3__check_button.raise_()
        self.s4__PH_btn.raise_()
        self.s4_dt_btn.raise_()
        self.s4_fish_btn.raise_()
        self.s4__TEMP_btn.raise_()
        self.s4__OXY_btn.raise_()
        self.formGroupBox.raise_()
        self.hex_receive.raise_()
        self.s2__clear_button.raise_()
        self.timer_send_cb.raise_()
        self.lineEdit_3.raise_()
        self.dw.raise_()

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.formGroupBox.setTitle(_translate("Form", "通信设置"))
        self.s1__lb_1.setText(_translate("Form", "串口检测："))
        self.s1__box_1.setText(_translate("Form", "检测串口"))
        self.s1__lb_2.setText(_translate("Form", "串口选择："))
        self.s1__lb_3.setText(_translate("Form", "波特率："))
        self.s1__box_3.setItemText(0, _translate("Form", "115200"))
        self.s1__box_3.setItemText(1, _translate("Form", "2400"))
        self.s1__box_3.setItemText(2, _translate("Form", "4800"))
        self.s1__box_3.setItemText(3, _translate("Form", "9600"))
        self.s1__box_3.setItemText(4, _translate("Form", "14400"))
        self.s1__box_3.setItemText(5, _translate("Form", "19200"))
        self.s1__box_3.setItemText(6, _translate("Form", "38400"))
        self.s1__box_3.setItemText(7, _translate("Form", "57600"))
        self.s1__box_3.setItemText(8, _translate("Form", "76800"))
        self.s1__box_3.setItemText(9, _translate("Form", "12800"))
        self.s1__box_3.setItemText(10, _translate("Form", "230400"))
        self.s1__box_3.setItemText(11, _translate("Form", "460800"))
        self.s1__lb_4.setText(_translate("Form", "数据位："))
        self.s1__box_4.setItemText(0, _translate("Form", "8"))
        self.s1__box_4.setItemText(1, _translate("Form", "7"))
        self.s1__box_4.setItemText(2, _translate("Form", "6"))
        self.s1__box_4.setItemText(3, _translate("Form", "5"))
        self.s1__lb_5.setText(_translate("Form", "校验位："))
        self.s1__box_5.setItemText(0, _translate("Form", "N"))
        self.open_button.setText(_translate("Form", "启动系统"))
        self.close_button.setText(_translate("Form", "系统关闭"))
        self.s1__lb_6.setText(_translate("Form", "停止位："))
        self.s1__box_6.setItemText(0, _translate("Form", "1"))
        self.verticalGroupBox_1.setTitle(_translate("Form", "接受区"))
        self.verticalGroupBox_2.setTitle(_translate("Form", "发送区"))
        self.s3__send_text.setHtml(_translate("form",'青鱼-3月龄-1-7.4-0.21-28-STM_ND'))
        self.s3__send_button.setText(_translate("Form", "发送"))
        self.s3__clear_button.setText(_translate("Form", "清除"))
        self.s3__check_button.setText(_translate("Form", "校验"))
        self.s4__OXY_btn.setText(_translate("Form", "溶氧量曲线"))
        self.s4__TEMP_btn.setText(_translate("Form", "温度曲线"))
        self.s4__PH_btn.setText(_translate("Form", "PH曲线"))
        self.s4_dt_btn.setText(_translate('Form','输入查询日期'))
        self.s4_fish_btn.setText(_translate('Form','输入查询鱼种'))
        self.s4__PH_lab.setText(_translate("Form","PH  值:"))
        self.s4__OXY_lab.setText(_translate("Form","溶氧量:"))
        self.s4__TEMP_lab.setText(_translate("Form","温  度:"))
        self.s4__PH_val.setText(_translate("Form","7.5"))
        self.s4__TEMP_val.setText(_translate("Form","30"))
        self.s4__OXY_val.setText(_translate("Form","6"))
        self.formGroupBox1.setTitle(_translate("Form", "系统状态"))
        self.label.setText(_translate("Form", "已接收："))
        self.label_2.setText(_translate("Form", "已发送："))
        self.hex_receive.setText(_translate("Form", "Hex接收"))
        self.s2__clear_button.setText(_translate("Form", "清除"))
        self.timer_send_cb.setText(_translate("Form", "定时发送"))
        self.lineEdit_3.setText(_translate("Form", "1000"))
        self.dw.setText(_translate("Form", "ms/次"))

class sqlExcels():

    def CreateFile(self,fishname):
        '''
        -->
        创建一个文件，建立一个分区，分区名为当日日期
        '''
        day_CF=day
        wb=openpyxl.Workbook(r'C:\WorkSpace\Study\VSCode\excel\{0}\{1}.xlsx'.format(fishname,day_CF))
        wb.create_sheet('Sheet1')
        ws=wb.active
        ws.append(['记录产生时间','终端编号','酸碱度','温度','溶氧量','指令码'])
        wb.save(r'C:\WorkSpace\Study\VSCode\excel\{0}\{1}.xlsx'.format(fishname,day_CF))
        wb.close()

    def LogData(self,FeedBackCode,Msg):
        '''
        -->
        将从下位机接收到的数据解码以后填充到单元格里
        '''
        day_LD=day
        num=Msg[-1]
        print(num)
        num=int(num)
        fish=int(Msg[-2])
        Type=fish_name[fish]
        path=r'C:\WorkSpace\Study\VSCode\excel\{0}\{1}.xlsx'.format(Type,day_LD)
        if os.path.exists(path):
            wb=openpyxl.load_workbook(path)
            ws=wb['Sheet1']
            if dt.minute%1==0:
                ws.cell(row=num,column=1,value=minute)
                ws.cell(row=num,column=2,value=Msg[0])
                ws.cell(row=num,column=3,value=Msg[1])
                ws.cell(row=num,column=4,value=Msg[2])
                ws.cell(row=num,column=5,value=Msg[3])
                ws.cell(row=num,column=6,value=FeedBackCode)
            wb.save(r'C:\WorkSpace\Study\VSCode\excel\{0}\{1}.xlsx'.format(Type,day_LD))
            wb.close()
        else:
            self.CreateFile(Type)

class PyQt5_Serial(QtWidgets.QWidget, Ui_Form,sqlExcels):
    def __init__(self):
        super(PyQt5_Serial, self).__init__()
        self.s4_fish='青鱼'
        self.s4_dt='2020-6-11'
        self.backCode='######'
        self.setupUi(self)
        self.init()
        self.connect_sql=None
        self.cursor_sql=None
        self.receive=['','']
        self.setWindowTitle("智能鱼塘养殖系统")
        self.ser = serial.Serial()
        self.port_check()
        self.s2__receive_text.setStyleSheet('font-size:24px;')

        # 接收数据和发送数据数目置零
        self.data_num_received = 0
        self.lineEdit.setText(str(self.data_num_received))
        self.data_num_sended = 0
        self.lineEdit_2.setText(str(self.data_num_sended))

    def init(self):
        # 串口检测按钮
        self.s1__box_1.clicked.connect(self.port_check)

        # 串口信息显示
        self.s1__box_2.currentTextChanged.connect(self.port_imf)

        # 打开串口按钮
        self.open_button.clicked.connect(self.port_open)

        # 关闭串口按钮
        self.close_button.clicked.connect(self.port_close)

        # 发送数据按钮
        self.s3__send_button.clicked.connect(self.data_send)

        #校验即将发送的数据
        self.s3__check_button.clicked.connect(self.generateMsg)

        #获取用户输入的查询日期
        self.s4_dt_btn.clicked.connect(self.get_date)

        #获取用户输入的查询日期
        self.s4_fish_btn.clicked.connect(self.get_fish)

        #显示PH变化曲线
        self.s4__PH_btn.clicked.connect(self.PHPlot)        

        #显示温度变化曲线
        self.s4__TEMP_btn.clicked.connect(self.TempPlot)

        #显示溶氧量变化曲线
        self.s4__OXY_btn.clicked.connect(self.OxyPlot)

        # 定时发送数据
        self.timer_send = QTimer()
        self.timer_send.timeout.connect(self.data_send)
        self.timer_send_cb.stateChanged.connect(self.data_send_timer)

        # 定时器接收数据
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.data_receive)

        # 清除发送窗口
        self.s3__clear_button.clicked.connect(self.send_data_clear)

        # 清除接收窗口
        self.s2__clear_button.clicked.connect(self.receive_data_clear)

    # 串口检测
    def port_check(self):
        # 检测所有存在的串口，将信息存储在字典中
        self.Com_Dict = {}
        port_list = list(serial.tools.list_ports.comports())
        self.s1__box_2.clear()
        for port in port_list:
            self.Com_Dict["%s" % port[0]] = "%s" % port[1]
            self.s1__box_2.addItem(port[0])
        if len(self.Com_Dict) == 0:
            self.state_label.setText(" 无串口")

    def get_date(self):
        day=QtWidgets.QInputDialog.getText(self, "日期输入", "输入日期", text="2020-6-11")
        day=str(day[0])
        self.s4_lin_day.setText(day)
        self.s4_dt=day
    
    def get_fish(self):
        day=QtWidgets.QInputDialog.getText(self, "鱼种输入", "输入鱼种", text="青鱼")
        day=str(day[0])
        self.s4_lin_fish.setText(day)
        self.s4_fish=day
    
    def updateMsg(self,data):
        self.s4__PH_val.setText(data[0])
        self.s4__TEMP_val.setText(data[1])
        self.s4__OXY_val.setText(data[2])

    # 串口信息
    def port_imf(self):
        # 显示选定的串口的详细信息
        imf_s = self.s1__box_2.currentText()
        if imf_s != "":
            self.state_label.setText(self.Com_Dict[self.s1__box_2.currentText()])

    # 打开串口
    def port_open(self):
        self.ser.port = self.s1__box_2.currentText()
        self.ser.baudrate = int(self.s1__box_3.currentText())
        self.ser.bytesize = int(self.s1__box_4.currentText())
        self.ser.stopbits = int(self.s1__box_6.currentText())
        self.ser.parity = self.s1__box_5.currentText()
        try:
            self.ser.open()
        except:
            QMessageBox.critical(self, "Port Error", "此串口不能被打开！")
            return None

        # 打开串口接收定时器，周期为10ms
        self.timer.start(10)

        if self.ser.isOpen():
            self.open_button.setEnabled(False)
            self.close_button.setEnabled(True)
            self.formGroupBox1.setTitle("系统状态（已开启）")

    # 关闭串口
    def port_close(self):
        self.timer.stop()
        self.timer_send.stop()
        try:
            self.ser.close()
        except:
            pass
        self.open_button.setEnabled(True)
        self.close_button.setEnabled(False)
        self.lineEdit_3.setEnabled(True)
        # 接收数据和发送数据数目置零
        self.data_num_received = 0
        self.lineEdit.setText(str(self.data_num_received))
        self.data_num_sended = 0
        self.lineEdit_2.setText(str(self.data_num_sended))
        self.formGroupBox1.setTitle("系统状态（已关闭）")

    # 发送数据
    def data_send(self):
        if self.ser.isOpen():
            input_s = self.s3__send_text.toPlainText()
            if input_s != "":# 非空字符串
                # ascii发送
                input_s = (input_s + '\r\n').encode('utf-8')
                num = self.ser.write(input_s)
                self.data_num_sended += num
                self.lineEdit_2.setText(str(self.data_num_sended))
        else:
            pass
    
    def link(self):
        try:
            self.connect_sql=pymysql.connect(host='172.16.77.7',db='pzz',user='pzz',password='1')
        except Exception as event:
            print(event)
        else:
            print('Link!')
            self.cursor_sql=self.connect_sql.cursor()

    def add(self,fish_name,device_name,tm,PHValue,OxyDegree,useyear,use_month,use_day,use_time):
        sql='insert into sb2 values(%s,%s,%s,%s,%s,%s,%s,%s,%s)'
        try:
            self.cursor_sql.execute(sql,(fish_name,device_name,tm,PHValue,OxyDegree,useyear,use_month,use_day,use_time))
        except Exception as ret:
            print(ret)
        else:
            self.connect_sql.commit()
            print('Success!')

    '青鱼-3月龄-1-7.4-0.21-28-STM_ND'
    def generateMsg(self):
        msg=[]
        text = self.s3__send_text.toPlainText()
        if text!='':
            text=text.split('-')
            print(text)
        else:
            print('the input is blank')
            return 0
        for index in fish_name.keys():
            if text[0]==fish_name[index]:
                msg.append(index)
            elif text[0] not in fish_name.values():
                print('wrong input one')
                return 0
        for index in fish_age.keys():
            if text[1]==fish_age[index]:
                msg.append(index)
            elif text[1] not in fish_age.values():
                    print('wrong input two')
                    return 1
        if text[2]=='1':
            msg.append(1)
            for index in pool_ph.keys():
                if text[3]==pool_ph[index]:

                    msg.append(index)
                elif text[3] not in pool_ph.values():
                    print('wrong input three')
                    return 3
            for index in pool_oxy.keys():
                if text[4]==pool_oxy[index]:

                    msg.append(index)
                elif text[4] not in pool_oxy.values():
                    print('wrong input four')
                    return 4
            for index in pool_temp.keys():
                if text[5]==pool_temp[index]:

                    msg.append(index)
                elif text[5] not in pool_temp.values():
                    print('wrong input five')
                    return 5
            msg.append(device[text[-1]])
        else:
            msg.extend([0,0,0,0])
            msg.append(device[text[-1]])
        msg=[str(i) for i in msg]
        msg=''.join(msg)
        self.backCode=msg#msg既为加密后的指令
        self.s3__send_text.setText(msg)

    # 接收数据
    def data_receive(self):
        try:
            num = self.ser.inWaiting()
        except:
            self.port_close()
            return None
        if num > 0:
            data = self.ser.read(num)
            num = len(data)
            # hex显示
            if self.hex_receive.checkState():
                out_s = ''
                for i in range(0, len(data)):
                    out_s = out_s + '{:02X}'.format(data[i]) + ' '
                self.receive_data_clear()
                self.s2__receive_text.insertPlainText(out_s)
            else:
                # 串口接收到的字符串为b'123',要转化成unicode字符串才能输出到窗口中去
                self.receive_data_clear()
                data_str=data.decode('utf-8')
                msg=data_str.split('-')#['STM_TWO', '7.44', '0.00', '49.09', '1', '102']
                print(msg)
                self.backCode='2114350'
                data=msg[2:5]
                if msg[0]=='STM_ONE':
                    self.receive[0]=data_str
                elif msg[0]=='STM_TWO':
                    self.receive[1]=data_str
                self.link()
                if msg[-2]=='1':
                    fuck_fish='Green'
                elif msg[-2]=='0':
                    fuck_fish='LuoFei'
                self.add(fuck_fish,msg[0],msg[2],msg[3],msg[1],dt.year,dt.month,dt.day,minute)
                self.s2__receive_text.insertPlainText(self.receive[0]+'\n')
                self.s2__receive_text.insertPlainText(self.receive[1])
                if self.backCode!='######':# and dt.minute==59:
                    self.LogData(FeedBackCode=self.backCode,Msg=msg)
                    self.updateMsg(data)

            # 统计接收字符的数量
            self.data_num_received += num
            self.lineEdit.setText(str(self.data_num_received))

            # 获取到text光标
            textCursor = self.s2__receive_text.textCursor()
            # 滚动到底部
            textCursor.movePosition(textCursor.End)
            # 设置光标到text中去
            self.s2__receive_text.setTextCursor(textCursor)
        else:
            pass

    #将多个数组融合为一个数组，将多个数组的平均值作为最终结果
    def mergeList(self,res1,res2):
        res=[]
        for i,j in zip(res1,res2):
            if i!=0 and j!=0:
                res.append((i+j)/2)

    # 定时发送数据
    def data_send_timer(self):
        if self.timer_send_cb.isChecked():
            self.timer_send.start(int(self.lineEdit_3.text()))
            self.lineEdit_3.setEnabled(False)
        else:
            self.timer_send.stop()
            self.lineEdit_3.setEnabled(True)

    # 清除显示
    def send_data_clear(self):
        self.s3__send_text.setText("")

    def receive_data_clear(self):
        self.s2__receive_text.setText("")

    def PHPlot(self):
        lists=[]
        name=self.s4_fish
        day_PP=self.s4_dt
        excel_path=r'C:\WorkSpace\Study\VSCode\excel\{0}\{1}.xlsx'.format(name,day_PP)
        self.data=xlrd.open_workbook(excel_path)  ##获取文本对象
        self.table=self.data.sheets()[0]   #根据index获取某个sheet
        self.rows=self.table.ncols  #3获取当前sheet页面的总行数,把每一行数据作为list放到 list
        for i in range(self.rows):
            col=self.table.col_values(i)  ##获取每一列数据
            lists.append(col[1:])
        time=lists[0]
        ph=lists[3]
        fig=plt.figure()
        ax=fig.add_subplot(1,1,1)
        ax.plot(time,ph)
        plt.show()

    def OxyPlot(self):
        lists=[]
        name=self.s4_fish
        day_QP=self.s4_dt
        excel_path=r'C:\WorkSpace\Study\VSCode\excel\{0}\{1}.xlsx'.format(name,day_QP)
        self.data=xlrd.open_workbook(excel_path)  ##获取文本对象
        self.table=self.data.sheets()[0]   #根据index获取某个sheet
        self.rows=self.table.ncols  #3获取当前sheet页面的总行数,把每一行数据作为list放到 list
        for i in range(self.rows):
            col=self.table.col_values(i)  ##获取每一列数据
            lists.append(col[1:])
        time=lists[0]
        oxy=lists[4]
        fig=plt.figure()
        ax=fig.add_subplot(1,1,1)
        ax.plot(time,oxy)
        plt.show()

    def TempPlot(self):
        lists=[]
        name=self.s4_fish
        day_TP=self.s4_dt
        excel_path=r'C:\WorkSpace\Study\VSCode\excel\{0}\{1}.xlsx'.format(name,day_TP)
        self.data=xlrd.open_workbook(excel_path)  ##获取文本对象
        self.table=self.data.sheets()[0]   #根据index获取某个sheet
        self.rows=self.table.ncols  #3获取当前sheet页面的总行数,把每一行数据作为list放到 list
        for i in range(self.rows):
            col=self.table.col_values(i)  ##获取每一列数据
            lists.append(col[1:])
        time=lists[0]
        temp=lists[2]
        fig=plt.figure()
        ax=fig.add_subplot(1,1,1)
        ax.plot(time,temp)
        plt.show()

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    myshow = PyQt5_Serial()
    myshow.show()
    sys.exit(app.exec_())
